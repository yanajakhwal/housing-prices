## EDA: Exploratory Data Analysis

import openpyxl
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl.workbook import Workbook
from io import BytesIO


def sale_price(df, name):
    # create histogram
    plt.figure(figsize=(10, 6))  # Set the figure size
    df['SalePrice'].hist(bins=30, color='blue', alpha=0.7)
    plt.title("Distribution of Sale Prices")
    plt.xlabel("Sale Price")
    plt.ylabel("Frequency")
    # plt.show() for interactive histogram

    # save plot to buffer
    buffer = BytesIO()
    plt.savefig(buffer, format='png')
    buffer.seek(0)  # reset buffer position

    # open workbook and create a new sheet for prices
    wb = openpyxl.load_workbook(f'data/eda/eda_{name}.xlsx')
    if "prices" not in wb.sheetnames:
        wb.create_sheet("prices")
    ws = wb["prices"]

    # insert the plot into the workbook
    img = openpyxl.drawing.image.Image(buffer)
    ws.add_image(img, 'A1')

    # save the workbook
    wb.save(f'data/eda/eda_{name}.xlsx')

    # clear the buffer
    buffer.close()


## general info about the data.
def gen_data(df, name):
    with pd.ExcelWriter(f'data/eda/eda_{name}.xlsx') as writer:

        # info sheet
        feature = df.columns
        nulls = df.isnull().sum().values
        non_nulls = len(df) - nulls
        type_ = df.dtypes.values
        info_df = pd.DataFrame({
            "feature": feature,
            "non-nulls": non_nulls,
            "nulls": nulls,
            "type": type_
        })
        info_df.to_excel(writer, sheet_name="info", index=False)

        # description sheet
        desc_df = df.describe()
        desc_df.insert(0, "Statistic", ["count", "mean", "std", "min", "25%", "50%", "75%", "max"])
        desc_df.to_excel(writer, sheet_name="description", index=False)
    

def main():
    train = pd.read_csv('data/raw/train.csv')
    test = pd.read_csv("data/raw/test.csv")

    combined = pd.concat([train, test], ignore_index=True)
    combined = combined.drop(columns = ['SalePrice'])

    gen_data(combined, "combined")
    sale_price(train, "combined")

if __name__ == '__main__':
    main()
    print("open data/eda_combined.xlsx to see a summary of the raw data!")